"""
Generate weather + holidays data for Lebanon 2025.
Fetches REAL weather from Open-Meteo (free, no API key needed).
Falls back to realistic climate averages if the API is unavailable.

Output: data/weather_holidays_2025.xlsx
"""
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Beirut coordinates
LAT = 33.888
LON = 35.495

# Fallback climate data (realistic Lebanon averages for Beirut)
FALLBACK_CLIMATE = [
    # month, avg_max, avg_min, avg_mean, rainfall_mm, rainy_days, hot_days, cold_days
    (1, 16.5, 9.1, 12.8, 165.0, 14, 0, 22),
    (2, 17.2, 9.5, 13.4, 140.0, 12, 0, 18),
    (3, 19.5, 11.2, 15.4, 95.0, 9, 0, 8),
    (4, 23.0, 14.0, 18.5, 45.0, 4, 2, 0),
    (5, 26.5, 17.5, 22.0, 15.0, 2, 8, 0),
    (6, 29.5, 21.0, 25.3, 2.0, 0, 20, 0),
    (7, 31.5, 23.5, 27.5, 1.0, 0, 28, 0),
    (8, 32.0, 24.0, 28.0, 1.5, 0, 29, 0),
    (9, 30.0, 22.0, 26.0, 5.0, 1, 22, 0),
    (10, 27.0, 18.5, 22.8, 35.0, 4, 12, 0),
    (11, 22.0, 14.5, 18.3, 90.0, 8, 3, 3),
    (12, 18.5, 11.0, 14.8, 150.0, 13, 0, 15),
]

LEBANESE_HOLIDAYS_2025 = {
    1: ["New Year"],
    2: ["Saint Maron"],
    3: ["Ramadan"],
    4: ["Easter", "Eid al-Fitr"],
    5: ["Labor Day", "Liberation Day"],
    6: ["Eid al-Adha"],
    7: ["Summer tourism peak"],
    8: ["Assumption Day", "Summer tourism peak"],
    9: ["School year start"],
    10: [],
    11: ["Independence Day"],
    12: ["Christmas", "Fireworks season"],
}


def fetch_weather_from_api():
    """Fetch real weather from Open-Meteo. Returns DataFrame or None if fails."""
    url = "https://archive-api.open-meteo.com/v1/archive"
    params = {
        "latitude": LAT,
        "longitude": LON,
        "start_date": "2025-01-01",
        "end_date": "2025-12-20",
        "daily": "temperature_2m_max,temperature_2m_min,temperature_2m_mean,precipitation_sum",
        "timezone": "Asia/Beirut"
    }
    print("Fetching weather from Open-Meteo for Beirut...")
    try:
        response = requests.get(url, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()

        df = pd.DataFrame({
            "date": pd.to_datetime(data["daily"]["time"]),
            "temp_max": data["daily"]["temperature_2m_max"],
            "temp_min": data["daily"]["temperature_2m_min"],
            "temp_mean": data["daily"]["temperature_2m_mean"],
            "precipitation": data["daily"]["precipitation_sum"],
        })

        df["month"] = df["date"].dt.month

        # Aggregate to monthly
        monthly = df.groupby("month").agg(
            avg_temp_max=("temp_max", "mean"),
            avg_temp_min=("temp_min", "mean"),
            avg_temp_mean=("temp_mean", "mean"),
            total_rainfall_mm=("precipitation", "sum"),
            rainy_days=("precipitation", lambda x: (x > 1).sum()),
            hot_days=("temp_max", lambda x: (x > 30).sum()),
            cold_days=("temp_min", lambda x: (x < 10).sum()),
        ).reset_index()

        # Round
        for col in ["avg_temp_max", "avg_temp_min", "avg_temp_mean", "total_rainfall_mm"]:
            monthly[col] = monthly[col].round(1)

        print(f"  ✓ Got real weather data for {len(monthly)} months")
        return monthly
    except Exception as e:
        print(f"  ⚠ API failed: {e}")
        print(f"  → Falling back to climate averages")
        return None


def build_fallback_weather():
    """Build weather DataFrame from hardcoded Lebanon climate averages."""
    rows = []
    for month, tmax, tmin, tmean, rain, rainy, hot, cold in FALLBACK_CLIMATE:
        rows.append({
            "month": month,
            "avg_temp_max": tmax,
            "avg_temp_min": tmin,
            "avg_temp_mean": tmean,
            "total_rainfall_mm": rain,
            "rainy_days": rainy,
            "hot_days": hot,
            "cold_days": cold,
        })
    return pd.DataFrame(rows)


def get_season(month):
    if month in [12, 1, 2]: return "Winter"
    if month in [3, 4, 5]: return "Spring"
    if month in [6, 7, 8]: return "Summer"
    return "Fall"


def add_holidays(weather_df):
    """Add holiday columns to weather DataFrame."""
    rows = []
    for _, row in weather_df.iterrows():
        m = int(row["month"])
        holidays = LEBANESE_HOLIDAYS_2025.get(m, [])
        rows.append({
            "season": get_season(m),
            "has_ramadan": int(m in [2, 3]),
            "has_eid": int(m in [4, 6]),
            "has_christmas_fireworks": int(m == 12),
            "is_summer_tourism_peak": int(m in [7, 8]),
            "has_school_year_start": int(m == 9),
            "holiday_names": "; ".join(holidays) if holidays else "—",
        })
    holiday_df = pd.DataFrame(rows)
    return pd.concat([weather_df.reset_index(drop=True), holiday_df], axis=1)


def save_to_excel(df):
    """Save combined data to Excel with nice formatting."""
    # Reorder columns
    cols_order = [
        "month", "season",
        "avg_temp_max", "avg_temp_min", "avg_temp_mean",
        "total_rainfall_mm", "rainy_days", "hot_days", "cold_days",
        "has_ramadan", "has_eid", "has_christmas_fireworks",
        "is_summer_tourism_peak", "has_school_year_start",
        "holiday_names"
    ]
    df = df[cols_order]

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Weather & Holidays 2025"

    # Header styling
    for col, name in enumerate(df.columns, start=1):
        cell = sheet.cell(row=1, column=col, value=name)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", start_color="C8102E")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_idx + 2, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center", horizontal="center")
            cell.border = border

    # Column widths
    widths = [8, 10, 14, 14, 14, 18, 12, 12, 12, 12, 10, 22, 22, 18, 35]
    for i, w in enumerate(widths, start=1):
        if i <= 26:
            sheet.column_dimensions[chr(64 + i)].width = w

    sheet.row_dimensions[1].height = 28
    sheet.freeze_panes = "A2"

    output_path = "data/weather_holidays_2025.xlsx"
    wb.save(output_path)
    print(f"\n✓ Saved {output_path}")


def main():
    print("=" * 60)
    print("Generating weather + holidays for Lebanon 2025")
    print("=" * 60)

    # Try real API first, fall back if needed
    weather_df = fetch_weather_from_api()
    if weather_df is None:
        weather_df = build_fallback_weather()
        print("  ✓ Using fallback climate data")

    # Add holiday columns
    combined = add_holidays(weather_df)

    # Show preview
    print("\nPreview:")
    print(combined[["month", "season", "avg_temp_mean", "total_rainfall_mm",
                    "rainy_days", "hot_days", "has_ramadan", "holiday_names"]].to_string(index=False))

    # Save
    save_to_excel(combined)

    print("\n✅ Done! Next step: run train_model_v2.py")


if __name__ == "__main__":
    main()